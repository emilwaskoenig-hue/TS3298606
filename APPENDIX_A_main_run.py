"""
ESG Pipeline — Full Run Script
================================
Three-part pipeline for processing all 3,038 proposals.

PART A: Agent 1 extraction (GPT-4o-mini)
  - Fetches HTML from SEC EDGAR, extracts proposal text
  - Runs live on your laptop (~45 min)
  - Output: main_run_a1_results.jsonl

PART B: Agent 2 batch submission (Claude Sonnet via Batch API)
  - Submits all scoring requests to Anthropic's Batch API
  - Takes ~2 minutes, then laptop can be closed
  - Output: main_run_batch_id.txt (save this — needed for Part C)

PART C: Batch results download (run next morning)
  - Downloads completed batch from Anthropic
  - Writes final CSV: main_run_output.csv
  - Takes ~2 minutes

USAGE:
  Part A:  python3 main_run.py --part a
  Part B:  python3 main_run.py --part b
  Part C:  python3 main_run.py --part c

SETUP:
  pip3 install openai anthropic openpyxl beautifulsoup4 certifi

API KEYS (set before each session):
  export OPENAI_API_KEY="sk-your-key"
  export ANTHROPIC_API_KEY="sk-ant-your-key"

ESTIMATED COST:
  Agent 1 (GPT-4o-mini): ~$4
  Agent 2 (Claude Sonnet Batch, 50% discount): ~$40
  Total: ~$44
"""

import os, sys, json, csv, time, urllib.request, ssl, re, threading
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

# ─── CONFIG ──────────────────────────────────────────────────────────────────
EXCEL_FILE      = "Full_dataset.xlsx"
SHEET_NAME      = 0
ID_COLUMN       = "ID_Fixed"
TITLE_COLUMN    = "Title"
LABEL_COLUMN    = "Proposal Type Specific"
URL_COLUMN      = "Link"

A1_RESULTS_FILE = "main_run_a1_results.jsonl"
BATCH_ID_FILE   = "main_run_batch_id.txt"
OUTPUT_CSV      = "main_run_output.csv"
LOG_FILE_A1     = "main_run_agent1_log.jsonl"
LOG_FILE_A2     = "main_run_agent2_log.jsonl"

A1_MODEL        = "gpt-4o-mini"
A2_MODEL        = "claude-sonnet-4-6"
A1_MAX_TOKENS   = 2000
A2_MAX_TOKENS   = 2000
PROPOSAL_WINDOW = 6000
MAX_WORKERS     = 5
MAX_A1_RETRIES  = 6

ADVISORY_VOTE_LABELS = {
    "say on pay", "advisory vote exec", "advisory vote on exec",
    "advisory vote on compensation", "say on frequency"
}
STRICT_KW = [
    "whereas", "resolved", "shareholders request", "shareholders urge",
    "shareholders recommend", "co-filers have advised",
    "intend to submit the following", "is hereby requested",
    "be it resolved", "will present the following proposal"
]

# ─── SSL ─────────────────────────────────────────────────────────────────────
def make_ssl_context():
    try:
        import certifi
        return ssl.create_default_context(cafile=certifi.where())
    except ImportError:
        pass
    try:
        return ssl.create_default_context()
    except Exception:
        return ssl._create_unverified_context()

SSL_CTX = make_ssl_context()

# ─── URL UTILITIES ────────────────────────────────────────────────────────────
def clean_url(url):
    if not url: return url
    url = str(url).strip()
    if "ix?doc=" in url:
        url = "https://www.sec.gov" + url.split("ix?doc=")[1]
    elif "ixviewer/ix.html?doc=" in url:
        url = "https://www.sec.gov" + url.split("?doc=")[1]
    return url

def split_url(url):
    if "#" in url:
        p = url.split("#", 1)
        return p[0], p[1]
    return url, None

# ─── HTML CACHE (thread-safe with per-URL locks) ──────────────────────────────
_html_cache = {}
_cache_lock = threading.Lock()
_url_locks = {}
_url_locks_lock = threading.Lock()

SEC_HEADERS = {
    "User-Agent": "Mozilla/5.0 Academic Research thesis@university.edu",
    "Accept": "text/html,application/xhtml+xml",
}

def fetch_html_cached(base_url, timeout=30):
    with _cache_lock:
        if base_url in _html_cache:
            cached = _html_cache[base_url]
            return (cached, None) if cached else (None, "Previously failed")
    with _url_locks_lock:
        if base_url not in _url_locks:
            _url_locks[base_url] = threading.Lock()
        url_lock = _url_locks[base_url]
    with url_lock:
        with _cache_lock:
            if base_url in _html_cache:
                cached = _html_cache[base_url]
                return (cached, None) if cached else (None, "Previously failed")
        try:
            req = urllib.request.Request(base_url, headers=SEC_HEADERS)
            with urllib.request.urlopen(req, timeout=timeout, context=SSL_CTX) as r:
                html = r.read().decode("utf-8", errors="replace")
            with _cache_lock:
                _html_cache[base_url] = html
            return html, None
        except ssl.SSLError:
            with _cache_lock:
                _html_cache[base_url] = None
            return None, "SSL_ERROR: run Install Certificates.command"
        except Exception as e:
            with _cache_lock:
                _html_cache[base_url] = None
            return None, f"{type(e).__name__}: {e}"

def is_edgar_index(html):
    return "Document Format Files" in html and "DEF 14A" in html

def extract_primary_url_from_index(html):
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "html.parser")
    for row in soup.find_all("tr"):
        cells = row.find_all("td")
        if len(cells) >= 4:
            if "DEF 14A" in cells[3].get_text(strip=True):
                link = row.find("a", href=True)
                if link:
                    href = link["href"]
                    return ("https://www.sec.gov" + href) if href.startswith("/") else href
    return None

# ─── TEXT EXTRACTION ──────────────────────────────────────────────────────────
def extract_around_fragment(html, fragment, window=PROPOSAL_WINDOW):
    from bs4 import BeautifulSoup
    try:
        soup = BeautifulSoup(html, "html.parser")
    except Exception:
        # Fallback: lxml is more tolerant of malformed HTML entities
        try:
            soup = BeautifulSoup(html, "lxml")
        except Exception:
            return None
    anchor = soup.find(id=fragment) or soup.find(attrs={"name": fragment})
    if not anchor:
        return None
    texts, chars = [], 0
    for elem in anchor.find_all_next(string=True):
        t = elem.strip()
        if t:
            texts.append(t)
            chars += len(t)
        if chars >= window:
            break
    return " ".join(texts) if texts else None

def is_advisory_vote(title, label):
    return any(m in (title + " " + label).lower() for m in ADVISORY_VOTE_LABELS)

def extract_advisory_vote(html, window=PROPOSAL_WINDOW):
    from bs4 import BeautifulSoup
    try:
        soup = BeautifulSoup(html, "html.parser")
    except Exception:
        try:
            soup = BeautifulSoup(html, "lxml")
        except Exception:
            return None, None
    text = soup.get_text(separator=" ", strip=True)
    advisory_kw = ["nonbinding", "non-binding", "advisory", "named executive officer compensation",
                   "executive compensation", "say on pay", "say-on-pay", "advisory resolution"]
    pos = 0
    while True:
        idx = text.lower().find("resolved", pos)
        if idx == -1: break
        ctx = (text[max(0, idx-1000):idx] + text[idx:idx+600]).lower()
        if any(k in ctx for k in advisory_kw):
            return text[max(0, idx-400):min(len(text), idx+window)], "advisory_vote_resolved"
        pos = idx + 1
    # Pass 2: filings that use header phrasing instead of RESOLVED
    for kw in ["approve, on a non-binding advisory basis", "approve, on an advisory basis",
               "approve the compensation of", "advisory shareholder vote"]:
        idx = text.lower().find(kw)
        if idx >= 0:
            return text[max(0, idx-400):min(len(text), idx+window)], "advisory_vote_header"
    # Pass 3: ALL-CAPS board recommendation used as resolution (e.g. Cardinal Health)
    for kw in ["NON-BINDING ADVISORY BASIS", "ADVISORY BASIS, OF THE COMPENSATION"]:
        idx = text.find(kw)
        if idx >= 0:
            return text[max(0, idx-400):min(len(text), idx+window)], "advisory_vote_boardrec"
    return None, None

def make_search_candidates(title):
    title_norm = title.replace('\u201c', '"').replace('\u201d', '"').replace('\u2018', "'").replace('\u2019', "'")
    title_bare = title_norm.replace('"', '').replace("'", "").strip()
    stop = {'report','on','a','the','of','and','for','to','in',
            'advisory','vote','exec.','comp.','assessment'}
    candidates = [title_norm, title_bare, title_bare.upper(), title_norm.replace("&", "and")]
    words = title_bare.split()
    if len(words) >= 4:
        candidates.append(" ".join(words[:4]))
    sig = [w for w in words if w.lower() not in stop and len(w) > 3]
    if sig:
        candidates.append(" ".join(sig[:3]))
        if len(sig) >= 2: candidates.append(sig[0])
        if len(sig) >= 3: candidates.append(sig[1])
    seen, out = set(), []
    for c in candidates:
        if c and c.lower() not in seen:
            seen.add(c.lower()); out.append(c)
    return out

def extract_by_title_search(html, title, window=PROPOSAL_WINDOW):
    from bs4 import BeautifulSoup
    try:
        soup = BeautifulSoup(html, "html.parser")
    except Exception:
        try:
            soup = BeautifulSoup(html, "lxml")
        except Exception:
            return None, None
    full = soup.get_text(separator=" ", strip=True)
    for term in make_search_candidates(title):
        pos = 0
        while True:
            idx = full.lower().find(term.lower(), pos)
            if idx == -1: break
            if any(kw in full[idx:idx+800].lower() for kw in STRICT_KW):
                return full[idx:min(len(full), idx+window)], f"title_search[{term[:30]}]"
            pos = idx + 1
    return None, None

def retrieve_text(url, title, label):
    base_url, fragment = split_url(url)
    html, err = fetch_html_cached(base_url)
    if not html:
        return None, None, "EXTRACTION_FAILURE", err
    if is_edgar_index(html):
        doc_url = extract_primary_url_from_index(html)
        if not doc_url:
            return None, None, "TEXT_MISSING", "No DEF14A link in index"
        html, err = fetch_html_cached(doc_url)
        if not html:
            return None, None, "EXTRACTION_FAILURE", err
        fragment = None
    if fragment:
        text = extract_around_fragment(html, fragment)
        if text and len(text) > 100:
            return text, "fragment_anchor", None, None
    if is_advisory_vote(title, label):
        text, hint = extract_advisory_vote(html)
        if text and len(text) > 100:
            return text, hint, None, None
    text, hint = extract_by_title_search(html, title)
    if text and len(text) > 100:
        return text, hint, None, None
    return None, None, "TEXT_MISSING", f"No match: {title[:50]}"

# ─── AGENT 1 PROMPT ───────────────────────────────────────────────────────────
A1_SYSTEM = """Extract a shareholder proposal verbatim from the SEC DEF 14A filing section provided.

Extract: proposal heading, RESOLVED clause, WHEREAS clauses, supporting statement.
Exclude: board recommendations, vote tallies, content after supporting statement.
If text exceeds ~1,100 words, summarise preserving ALL ESG-relevant content including annexes and cross-references.

BOARD RECOMMENDATION: Return true (FOR), false (AGAINST), or null. Direction only.

IMPORTANT: Return extracted_text as a single flat string, not nested JSON.

Return JSON only:
Success: {"extracted_text":"...","extraction_method":"raw","board_support":true|false|null,"null_flag":false,"null_reason":null}
Failure: {"extracted_text":null,"extraction_method":null,"board_support":null,"null_flag":true,"null_reason":"TEXT_MISSING"}"""

def get_retry_wait(err_msg):
    m = re.search(r'try again in (\d+(?:\.\d+)?)(ms|s)', str(err_msg))
    if m:
        val, unit = float(m.group(1)), m.group(2)
        return (val/1000 if unit == "ms" else val) + 0.5
    return 5.0

def call_a1(client, title, label, text):
    from openai import RateLimitError
    user_msg = f"Extract this proposal.\nTitle: {title}\nLabel: {label}\n\n--- DOCUMENT SECTION ---\n{text}"
    for attempt in range(MAX_A1_RETRIES):
        try:
            resp = client.chat.completions.create(
                model=A1_MODEL, max_tokens=A1_MAX_TOKENS,
                messages=[{"role": "system", "content": A1_SYSTEM},
                          {"role": "user", "content": user_msg}])
            raw = resp.choices[0].message.content
            t = raw.strip()
            if t.startswith("```"):
                lines = t.split("\n")
                t = "\n".join(lines[1:-1]) if lines[-1].strip() == "```" else "\n".join(lines[1:])
            return json.loads(t), raw
        except RateLimitError as e:
            wait = get_retry_wait(str(e))
            if attempt < MAX_A1_RETRIES - 1:
                time.sleep(wait)
            else:
                return None, f"RATE_LIMIT:{e}"
        except Exception as e:
            if attempt < MAX_A1_RETRIES - 1:
                time.sleep(2 ** attempt)
            else:
                return None, f"ERROR:{e}"
    return None, "MAX_RETRIES"

# ─── AGENT 2 PROMPT ───────────────────────────────────────────────────────────
A2_FEW_SHOT = """E1 Human Rights(S-only): {"esg_relevant":"YES","esg_confidence":95,"esg_reasoning":"Addresses human rights impacts in operations and value chain — substantive social issue.","e_active":0,"s_active":1,"g_active":0,"e_direction":null,"e_direction_reasoning":null,"e_direction_confidence":null,"e_specificity":null,"e_specificity_reasoning":null,"e_specificity_confidence":null,"s_direction":8,"s_direction_reasoning":"Requests human rights impact assessment across operations and value chain.","s_direction_confidence":93,"s_specificity":7,"s_specificity_reasoning":"Defined subject matter and scope; no quantitative targets or enforcement mechanisms.","s_specificity_confidence":90,"g_direction":null,"g_direction_reasoning":null,"g_direction_confidence":null,"g_specificity":null,"g_specificity_reasoning":null,"g_specificity_confidence":null,"low_confidence_flag":false,"mixed_direction_flag":false}
E2 Chairman Independence(G-only,dir=9): {"esg_relevant":"YES","esg_confidence":97,"esg_reasoning":"Board leadership structure and director independence are core governance matters.","e_active":0,"s_active":0,"g_active":1,"e_direction":null,"e_direction_reasoning":null,"e_direction_confidence":null,"e_specificity":null,"e_specificity_reasoning":null,"e_specificity_confidence":null,"s_direction":null,"s_direction_reasoning":null,"s_direction_confidence":null,"s_specificity":null,"s_specificity_reasoning":null,"s_specificity_confidence":null,"g_direction":9,"g_direction_reasoning":"Requires independent board chair whenever possible, separating chair and CEO roles.","g_direction_confidence":96,"g_specificity":9,"g_specificity_reasoning":"Specific policy, bylaw amendments if needed, defined implementation trigger.","g_specificity_confidence":95,"low_confidence_flag":false,"mixed_direction_flag":false}
E3 Say on Pay(G-only,dir=5): {"esg_relevant":"YES","esg_confidence":86,"esg_reasoning":"Say-on-pay advisory vote is a governance matter involving executive compensation oversight and shareholder rights.","e_active":0,"s_active":0,"g_active":1,"e_direction":null,"e_direction_reasoning":null,"e_direction_confidence":null,"e_specificity":null,"e_specificity_reasoning":null,"e_specificity_confidence":null,"s_direction":null,"s_direction_reasoning":null,"s_direction_confidence":null,"s_specificity":null,"s_specificity_reasoning":null,"s_specificity_confidence":null,"g_direction":5,"g_direction_reasoning":"Routine advisory vote on disclosed compensation — directionally neutral, not a governance reform.","g_direction_confidence":84,"g_specificity":7,"g_specificity_reasoning":"Shareholders vote to approve a defined compensation package on an advisory basis.","g_specificity_confidence":83,"low_confidence_flag":false,"mixed_direction_flag":false}
E4 Climate(E-only): {"esg_relevant":"YES","esg_confidence":95,"esg_reasoning":"Focuses on methane emissions and energy-sector client engagement — substantively environmental.","e_active":1,"s_active":0,"g_active":0,"e_direction":7,"e_direction_reasoning":"Mildly to moderately pro-environment: seeks methane reporting rather than direct emissions targets.","e_direction_confidence":92,"e_specificity":7,"e_specificity_reasoning":"Asks for a methane report with outlined topics; major design choices left to board.","e_specificity_confidence":89,"s_direction":null,"s_direction_reasoning":null,"s_direction_confidence":null,"s_specificity":null,"s_specificity_reasoning":null,"s_specificity_confidence":null,"g_direction":null,"g_direction_reasoning":null,"g_direction_confidence":null,"g_specificity":null,"g_specificity_reasoning":null,"g_specificity_confidence":null,"low_confidence_flag":false,"mixed_direction_flag":false}
E5 Lobbying(E+S+G): {"esg_relevant":"YES","esg_confidence":90,"esg_reasoning":"Political/lobbying expenditure alignment with stated values — governance-primary with E and S through cited policy areas.","e_active":1,"s_active":1,"g_active":1,"e_direction":6,"e_direction_reasoning":"Part of alignment test concerns climate commitments and political spending.","e_direction_confidence":84,"e_specificity":null,"e_specificity_reasoning":null,"e_specificity_confidence":null,"s_direction":7,"s_direction_reasoning":"Alignment framed around health affordability and social equity.","s_direction_confidence":86,"s_specificity":null,"s_specificity_reasoning":null,"s_specificity_confidence":null,"g_direction":8,"g_direction_reasoning":"Transparency over corporate political spending is itself the primary governance subject.","g_direction_confidence":89,"g_specificity":8,"g_specificity_reasoning":"Annual report, specified contents, explanations of incongruent expenditures, follow-up actions.","g_specificity_confidence":88,"low_confidence_flag":false,"mixed_direction_flag":false}
E6 Golden Parachutes(G-only,dir=9): {"esg_relevant":"YES","esg_confidence":96,"esg_reasoning":"Severance limits and shareholder approval of executive pay packages — squarely governance.","e_active":0,"s_active":0,"g_active":1,"e_direction":null,"e_direction_reasoning":null,"e_direction_confidence":null,"e_specificity":null,"e_specificity_reasoning":null,"e_specificity_confidence":null,"s_direction":null,"s_direction_reasoning":null,"s_direction_confidence":null,"s_specificity":null,"s_specificity_reasoning":null,"s_specificity_confidence":null,"g_direction":9,"g_direction_reasoning":"Tightens accountability around large severance and requires shareholder approval above a defined threshold.","g_direction_confidence":95,"g_specificity":9,"g_specificity_reasoning":"2.99x threshold defined, covered payments specified, approval trigger stated.","g_specificity_confidence":94,"low_confidence_flag":false,"mixed_direction_flag":false}
E7 Pay Gap(S-only): {"esg_relevant":"YES","esg_confidence":96,"esg_reasoning":"Requests reporting on racial and gender pay gaps and associated risks — substantive social issue.","e_active":0,"s_active":1,"g_active":0,"e_direction":null,"e_direction_reasoning":null,"e_direction_confidence":null,"e_specificity":null,"e_specificity_reasoning":null,"e_specificity_confidence":null,"s_direction":8,"s_direction_reasoning":"Pushes for pay gap transparency and workforce equity risk disclosure.","s_direction_confidence":95,"s_specificity":8,"s_specificity_reasoning":"Specifies report, metrics, and risk categories; no direct remediation targets.","s_specificity_confidence":92,"g_direction":null,"g_direction_reasoning":null,"g_direction_confidence":null,"g_specificity":null,"g_specificity_reasoning":null,"g_specificity_confidence":null,"low_confidence_flag":false,"mixed_direction_flag":false}
E8 Special Meetings(G-only,structural): {"esg_relevant":"YES","esg_confidence":97,"esg_reasoning":"Lowering the threshold for calling special shareholder meetings directly concerns shareholder rights and governance.","e_active":0,"s_active":0,"g_active":1,"e_direction":null,"e_direction_reasoning":null,"e_direction_confidence":null,"e_specificity":null,"e_specificity_reasoning":null,"e_specificity_confidence":null,"s_direction":null,"s_direction_reasoning":null,"s_direction_confidence":null,"s_specificity":null,"s_specificity_reasoning":null,"s_specificity_confidence":null,"g_direction":9,"g_direction_reasoning":"Expands shareholder rights to call a special meeting at a lower ownership threshold.","g_direction_confidence":96,"g_specificity":8,"g_specificity_reasoning":"Specifies threshold and governing documents to amend; exact drafting left to board.","g_specificity_confidence":92,"low_confidence_flag":false,"mixed_direction_flag":false}
E9 Whistleblower(S+G): {"esg_relevant":"YES","esg_confidence":94,"esg_reasoning":"Directly addresses non-discrimination, anti-harassment, and whistleblower protection through ethics code amendments.","e_active":0,"s_active":1,"g_active":1,"e_direction":null,"e_direction_reasoning":null,"e_direction_confidence":null,"e_specificity":null,"e_specificity_reasoning":null,"e_specificity_confidence":null,"s_direction":9,"s_direction_reasoning":"Strengthens protections against discrimination, harassment, and retaliation.","s_direction_confidence":94,"s_specificity":8,"s_specificity_reasoning":"Specifies policy areas and expected content in the revised code.","s_specificity_confidence":91,"g_direction":8,"g_direction_reasoning":"Amends a named governance document (Code of Ethics) — governance is co-primary to the social ask.","g_direction_confidence":88,"g_specificity":7,"g_specificity_reasoning":"Concrete amendment to a named governance document; final drafting left to the board.","g_specificity_confidence":85,"low_confidence_flag":false,"mixed_direction_flag":false}
E10 Civil Rights Audit(S-only): {"esg_relevant":"YES","esg_confidence":96,"esg_reasoning":"Requests an independent racial equity audit of impacts on nonwhite stakeholders and communities.","e_active":0,"s_active":1,"g_active":0,"e_direction":null,"e_direction_reasoning":null,"e_direction_confidence":null,"e_specificity":null,"e_specificity_reasoning":null,"e_specificity_confidence":null,"s_direction":9,"s_direction_reasoning":"Strongly advances social responsibility by seeking an independent racial equity audit for affected communities.","s_direction_confidence":95,"s_specificity":8,"s_specificity_reasoning":"Clear and concrete: independent third-party audit with stakeholder input; no quantified targets.","s_specificity_confidence":92,"g_direction":null,"g_direction_reasoning":null,"g_direction_confidence":null,"g_specificity":null,"g_specificity_reasoning":null,"g_specificity_confidence":null,"low_confidence_flag":false,"mixed_direction_flag":false}"""

A2_SYSTEM = f"""You are an ESG scoring assistant classifying shareholder proposals from SEC DEF 14A proxy filings.

STAGE 1 — ESG RELEVANCE GATE
YES if the proposal substantively engages E, S, or G (including anti-ESG). NO if ESG is merely contextual or incidental.
Return esg_relevant, esg_confidence (0-100), esg_reasoning (1-2 sentences). If NO, all dimension fields = null.

STAGE 2 — ACTIVE FLAGS AND SCORING (ESG-relevant only)
Determine e_active, s_active, g_active (1 or 0).

GOVERNANCE ACTIVATION BOUNDARY:
G is active ONLY when governance is a PRIMARY OR CO-PRIMARY subject: board composition, executive compensation, shareholder voting rights, director independence, anti-takeover provisions, oversight structures, bylaws.
G is NOT activated because the board receives a report request. A human rights, climate, AI, or social proposal is S/E-active even if it asks the board to commission a report.
Test: Would a governance specialist reading only the resolved clause identify a substantive governance reform? If no — G is not active.

DIRECTION (0-10): 10=strongly pro | 8-9=clearly pro | 6-8=disclosure requests | 5=neutral/routine | 2-4=anti-ESG | 0-1=strongly anti
5 = active-but-neutral. Do NOT use 5 for inactivity — inactivity is active=0 and null scores.

SPECIFICITY (0-10 or null): identifiable ask = requests/requires/urges concrete action, report, policy, target, disclosure. Commentary alone is not an ask.
9-10=defined targets/timelines | 7-8=clear ask, some discretion | 5-6=identifiable but vague | 3-4=aspirational | null=no identifiable ask

CONFIDENCE: 80-100=High | 50-79=Moderate | 0-49=Low
LOW_CONFIDENCE_FLAG: true if any applicable confidence < 80 (null confidences excluded)
MIXED_DIRECTION_FLAG: true if ≥2 active dims with one >5 and another <5

Return valid JSON only. Use null (not NA) for inactive fields.

Schema: {{"esg_relevant":"YES"|"NO","esg_confidence":int,"esg_reasoning":"...","e_active":0|1,"s_active":0|1,"g_active":0|1,"e_direction":int|null,"e_direction_reasoning":str|null,"e_direction_confidence":int|null,"e_specificity":int|null,"e_specificity_reasoning":str|null,"e_specificity_confidence":int|null,"s_direction":int|null,"s_direction_reasoning":str|null,"s_direction_confidence":int|null,"s_specificity":int|null,"s_specificity_reasoning":str|null,"s_specificity_confidence":int|null,"g_direction":int|null,"g_direction_reasoning":str|null,"g_direction_confidence":int|null,"g_specificity":int|null,"g_specificity_reasoning":str|null,"g_specificity_confidence":int|null,"low_confidence_flag":bool,"mixed_direction_flag":bool}}

FEW-SHOT EXAMPLES:
{A2_FEW_SHOT}"""

# ─── LOAD EXCEL ───────────────────────────────────────────────────────────────
def load_all_rows():
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.worksheets[SHEET_NAME]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    def ci(name):
        try: return headers.index(name)
        except ValueError: raise ValueError(f"Column '{name}' not found. Available: {headers}")
    id_i=ci(ID_COLUMN); ti=ci(TITLE_COLUMN); li=ci(LABEL_COLUMN); ui=ci(URL_COLUMN)
    rows = []
    for r in range(2, ws.max_row + 1):
        raw_url = ws.cell(r, ui + 1).value
        pid = ws.cell(r, id_i + 1).value
        if not pid: continue
        rows.append({
            "id": pid,
            "title": ws.cell(r, ti + 1).value or "",
            "label": ws.cell(r, li + 1).value or "",
            "url_original": raw_url,
            "url": clean_url(str(raw_url)) if raw_url else None
        })
    return rows

# ─── PART A: AGENT 1 EXTRACTION ──────────────────────────────────────────────
def run_part_a():
    from openai import OpenAI
    a1_key = os.environ.get("OPENAI_API_KEY")
    if not a1_key:
        print('\nOPENAI_API_KEY not set. Run: export OPENAI_API_KEY="sk-..."'); return

    client = OpenAI(api_key=a1_key)
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    t_start = time.time()

    print(f"\nPART A — Agent 1 Extraction | Run {run_id}")
    rows = load_all_rows()
    unique_bases = len(set(split_url(r["url"])[0] for r in rows if r["url"]))
    print(f"Loaded {len(rows)} proposals across {unique_bases} unique filing URLs.\n")

    # Load checkpoint if exists (allows resuming after crash)
    checkpoint_file = A1_RESULTS_FILE + ".checkpoint"
    completed_ids = set()
    if os.path.exists(checkpoint_file):
        with open(checkpoint_file, encoding="utf-8") as cf:
            for line in cf:
                try:
                    rec = json.loads(line)
                    completed_ids.add(str(rec["proposal_id"]))
                except Exception:
                    pass
        if completed_ids:
            print(f"  Resuming from checkpoint: {len(completed_ids)} proposals already fetched")

    print("Phase 1 — HTML fetch (sequential, cached):")
    fetched = []
    for i, row in enumerate(rows, 1):
        url, title, label = row["url"] or "", row["title"], row["label"]
        base, _ = split_url(url) if url else ("", "")
        tag = "C" if base in _html_cache else "F"
        if not url:
            fetched.append((row, None, "none", "EXTRACTION_FAILURE", "No URL"))
            if i % 100 == 0: print(f"  [{i}/{len(rows)}] {tag} NO_URL")
            continue
        text, hint, ec, em = retrieve_text(url, title, label)
        if i % 50 == 0 or not text:
            status = f"{hint[:20]}|{len(text):,}ch" if text else f"FAIL:{ec}"
            print(f"  [{i:4d}/{len(rows)}] {tag} {title[:35]} -> {status}")
        fetched.append((row, text, hint, ec or "", em or ""))

    saved = sum(1 for r in rows if r["url"]) - unique_bases
    to_process = sum(1 for _, t, *_ in fetched if t)
    print(f"\n  {unique_bases} fetches, {saved} cache hits | {to_process}/{len(rows)} ready for GPT-4o-mini\n")

    print(f"Phase 2 — GPT-4o-mini extraction ({MAX_WORKERS} workers):")
    results = {}
    results_lock = threading.Lock()
    print_lock = threading.Lock()

    def process(idx, row, py_text, hint):
        pid, title, label = row["id"], row["title"], row["label"]
        if not py_text:
            a1_result, a1_raw, sent = None, "NO_TEXT", 0
        else:
            a1_result, a1_raw = call_a1(client, title, label, py_text)
            sent = len(py_text)
            if isinstance(a1_raw, str) and any(a1_raw.startswith(p) for p in
                                                ["RATE_LIMIT","ERROR","MAX","JSON"]):
                a1_result = None

        with results_lock:
            results[idx] = (row, hint, a1_result, a1_raw, sent, py_text)

        if a1_result and not a1_result.get("null_flag"):
            tlen = len(a1_result.get("extracted_text") or "")
            board = a1_result.get("board_support")
            bs = "FOR" if board is True else "AGAINST" if board is False else "null"
            with print_lock:
                print(f"  ID={pid} OK[{a1_result.get('extraction_method','?')}] board={bs} {tlen}ch")
        else:
            reason = (a1_result or {}).get("null_reason", str(a1_raw)[:40])
            with print_lock:
                print(f"  ID={pid} FAIL: {reason}")

    args = [(i, row, text, hint) for i, (row, text, hint, *_) in enumerate(fetched)]
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futs = [ex.submit(process, *a) for a in args]
        for f in as_completed(futs):
            try: f.result()
            except Exception as e: print(f"  Thread error: {e}")

    # Write A1 results file
    a1_ok = 0
    with open(A1_RESULTS_FILE, "w", encoding="utf-8") as f, \
         open(LOG_FILE_A1, "w", encoding="utf-8") as logf:
        for i in range(len(fetched)):
            if i not in results: continue
            row, hint, a1_result, a1_raw, sent, py_text = results[i]
            a1_null = (a1_result or {}).get("null_flag", True) if a1_result else True
            if not a1_null: a1_ok += 1
            record = {
                "proposal_id": row["id"],
                "title": row["title"],
                "label": row["label"],
                "url_original": str(row["url_original"] or ""),
                "url_cleaned": row["url"] or "",
                "retrieval_method": hint,
                "a1_null_flag": a1_null,
                "a1_null_reason": (a1_result or {}).get("null_reason", fetched[i][3]),
                "extraction_method": (a1_result or {}).get("extraction_method", ""),
                "board_support": (a1_result or {}).get("board_support"),
                "extracted_text": (a1_result or {}).get("extracted_text") or "",
                "text_chars": len((a1_result or {}).get("extracted_text") or ""),
                "agent1_model": A1_MODEL,
                "run_id": run_id,
            }
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
            logf.write(json.dumps({
                "run_id": run_id, "proposal_id": row["id"],
                "url_cleaned": row["url"], "retrieval_method": hint,
                "text_sent_chars": sent, "model": A1_MODEL,
                "raw_response": str(a1_raw), "parsed_result": a1_result,
                "timestamp": datetime.now().isoformat()
            }, ensure_ascii=False) + "\n")

    elapsed = time.time() - t_start
    print(f"\n{'─'*55}")
    print(f"Part A done in {elapsed:.0f}s ({elapsed/len(rows):.1f}s/proposal)")
    print(f"Successful extractions: {a1_ok}/{len(rows)}")
    print(f"Results written to: {A1_RESULTS_FILE}")
    print(f"\nNext step: python3 main_run.py --part b")

# ─── PART B: BATCH SUBMISSION ─────────────────────────────────────────────────
def run_part_b():
    import anthropic
    a2_key = os.environ.get("ANTHROPIC_API_KEY")
    if not a2_key:
        print('\nANTHROPIC_API_KEY not set. Run: export ANTHROPIC_API_KEY="sk-ant-..."'); return

    if not os.path.exists(A1_RESULTS_FILE):
        print(f"A1 results file not found: {A1_RESULTS_FILE}")
        print("Run Part A first: python3 main_run.py --part a"); return

    client = anthropic.Anthropic(api_key=a2_key)

    print(f"\nPART B — Agent 2 Batch Submission")
    print(f"Reading Agent 1 results from {A1_RESULTS_FILE}...")

    with open(A1_RESULTS_FILE, encoding="utf-8") as f:
        a1_records = [json.loads(l) for l in f if l.strip()]

    # Build batch requests — only for proposals with extracted text
    requests_list = []
    skipped = 0
    for rec in a1_records:
        if rec.get("a1_null_flag") or not rec.get("extracted_text"):
            skipped += 1
            continue
        user_msg = (f"Score this proposal.\n\nTitle: {rec['title']}\n"
                    f"Label: {rec['label']}\n\n--- PROPOSAL TEXT ---\n{rec['extracted_text']}")
        requests_list.append({
            "custom_id": str(rec["proposal_id"]),
            "params": {
                "model": A2_MODEL,
                "max_tokens": A2_MAX_TOKENS,
                "system": A2_SYSTEM,
                "messages": [{"role": "user", "content": user_msg}]
            }
        })

    print(f"Batch size: {len(requests_list)} proposals ({skipped} skipped — null/failed A1)")
    print(f"Estimated cost: ~${len(requests_list) * 0.013:.0f} (Batch API 50% discount applied)")

    # Submit batch
    print("\nSubmitting batch to Anthropic... ", end="", flush=True)
    batch = client.beta.messages.batches.create(requests=requests_list)
    batch_id = batch.id
    print(f"Done.")
    print(f"\nBatch ID: {batch_id}")
    print(f"Status: {batch.processing_status}")

    with open(BATCH_ID_FILE, "w") as f:
        f.write(json.dumps({
            "batch_id": batch_id,
            "submitted_at": datetime.now().isoformat(),
            "request_count": len(requests_list),
            "model": A2_MODEL,
            "a1_results_file": A1_RESULTS_FILE,
        }, indent=2))

    print(f"\nBatch ID saved to: {BATCH_ID_FILE}")
    print("\n" + "="*55)
    print("YOU CAN NOW CLOSE YOUR LAPTOP.")
    print("Anthropic will process the batch over the next few hours.")
    print("Check status at: console.anthropic.com → Batches")
    print(f"\nWhen complete, run: python3 main_run.py --part c")
    print("="*55)

# ─── PART C: BATCH DOWNLOAD ───────────────────────────────────────────────────
def run_part_c():
    import anthropic
    a2_key = os.environ.get("ANTHROPIC_API_KEY")
    if not a2_key:
        print('\nANTHROPIC_API_KEY not set. Run: export ANTHROPIC_API_KEY="sk-ant-..."'); return

    if not os.path.exists(BATCH_ID_FILE):
        print(f"Batch ID file not found: {BATCH_ID_FILE}")
        print("Run Part B first: python3 main_run.py --part b"); return

    if not os.path.exists(A1_RESULTS_FILE):
        print(f"A1 results file not found: {A1_RESULTS_FILE}"); return

    with open(BATCH_ID_FILE) as f:
        batch_meta = json.load(f)
    batch_id = batch_meta["batch_id"]

    client = anthropic.Anthropic(api_key=a2_key)

    print(f"\nPART C — Batch Results Download")
    print(f"Batch ID: {batch_id}")

    batch = client.beta.messages.batches.retrieve(batch_id)
    print(f"Status: {batch.processing_status}")

    if batch.processing_status != "ended":
        print(f"\nBatch not yet complete (status: {batch.processing_status}).")
        print("Check again later: python3 main_run.py --part c")
        counts = batch.request_counts
        print(f"Progress: {counts.succeeded} succeeded, {counts.processing} processing, "
              f"{counts.errored} errored")
        return

    print(f"Batch complete. Downloading results...")

    # Load A1 records for merging
    with open(A1_RESULTS_FILE, encoding="utf-8") as f:
        a1_records = {str(rec["proposal_id"]): rec
                      for rec in (json.loads(l) for l in f if l.strip())}

    # Download and parse batch results
    a2_results = {}
    with open(LOG_FILE_A2, "w", encoding="utf-8") as logf:
        for result in client.beta.messages.batches.results(batch_id):
            custom_id = result.custom_id
            if result.result.type == "succeeded":
                raw = result.result.message.content[0].text
                t = raw.strip()
                if t.startswith("```"):
                    lines = t.split("\n")
                    t = "\n".join(lines[1:-1]) if lines[-1].strip() == "```" else "\n".join(lines[1:])
                try:
                    parsed = json.loads(t)
                    a2_results[custom_id] = parsed
                    logf.write(json.dumps({
                        "proposal_id": custom_id, "model": A2_MODEL,
                        "raw_response": raw, "parsed_result": parsed,
                        "timestamp": datetime.now().isoformat()
                    }, ensure_ascii=False) + "\n")
                except json.JSONDecodeError as e:
                    a2_results[custom_id] = {"parse_error": str(e)}
                    logf.write(json.dumps({"proposal_id": custom_id,
                        "parse_error": str(e), "raw": raw[:200]}) + "\n")
            else:
                a2_results[custom_id] = {"api_error": str(result.result.type)}

    print(f"Downloaded {len(a2_results)} results.")

    # Write final CSV
    csv_fields = [
        "proposal_id", "title", "existing_label", "url_original", "url_cleaned",
        "retrieval_method", "a1_null_flag", "a1_null_reason", "extraction_method",
        "board_support", "text_chars", "extracted_text",
        "esg_relevant", "esg_confidence", "esg_reasoning",
        "e_active", "s_active", "g_active",
        "e_direction", "e_direction_reasoning", "e_direction_confidence",
        "e_specificity", "e_specificity_reasoning", "e_specificity_confidence",
        "s_direction", "s_direction_reasoning", "s_direction_confidence",
        "s_specificity", "s_specificity_reasoning", "s_specificity_confidence",
        "g_direction", "g_direction_reasoning", "g_direction_confidence",
        "g_specificity", "g_specificity_reasoning", "g_specificity_confidence",
        "low_confidence_flag", "mixed_direction_flag",
        "agent1_model", "agent2_model", "parse_error"
    ]

    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as csvf:
        writer = csv.DictWriter(csvf, fieldnames=csv_fields)
        writer.writeheader()
        for pid_str, a1_rec in a1_records.items():
            a2 = a2_results.get(pid_str, {})
            row = {
                "proposal_id": a1_rec["proposal_id"],
                "title": a1_rec["title"],
                "existing_label": a1_rec["label"],
                "url_original": a1_rec["url_original"],
                "url_cleaned": a1_rec["url_cleaned"],
                "retrieval_method": a1_rec["retrieval_method"],
                "a1_null_flag": a1_rec["a1_null_flag"],
                "a1_null_reason": a1_rec["a1_null_reason"],
                "extraction_method": a1_rec["extraction_method"],
                "board_support": a1_rec["board_support"],
                "text_chars": a1_rec["text_chars"],
                "extracted_text": a1_rec["extracted_text"],
                "agent1_model": A1_MODEL,
                "agent2_model": A2_MODEL,
                "parse_error": a2.get("parse_error", a2.get("api_error", "")),
            }
            for field in csv_fields:
                if field not in row:
                    row[field] = a2.get(field, "")
            writer.writerow(row)

    a2_ok = sum(1 for v in a2_results.values() if "esg_relevant" in v)
    a2_err = len(a2_results) - a2_ok
    print(f"\nFinal CSV written: {OUTPUT_CSV}")
    print(f"A2 scored: {a2_ok}/{len(a2_results)} | Errors: {a2_err}")
    print(f"Null-flagged (A1 failures, not scored): {len(a1_records) - len(a2_results)}")

    # Summary stats
    scored = [v for v in a2_results.values() if "esg_relevant" in v]
    if scored:
        esg_yes = sum(1 for v in scored if v.get("esg_relevant") == "YES")
        esg_no = sum(1 for v in scored if v.get("esg_relevant") == "NO")
        lc = sum(1 for v in scored if v.get("low_confidence_flag"))
        md = sum(1 for v in scored if v.get("mixed_direction_flag"))
        print(f"\nOverall: ESG YES={esg_yes} NO={esg_no} | LC flags={lc} MD flags={md}")

    print(f"\nLogs: {LOG_FILE_A1} | {LOG_FILE_A2}")
    print("\nNext step: open main_run_output.csv in Excel for post-pipeline aggregation.")

# ─── ENTRY POINT ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 3 or sys.argv[1] != "--part":
        print(__doc__)
        print("\nUsage:")
        print("  python3 main_run.py --part a   # Extract text (run live, ~45 min)")
        print("  python3 main_run.py --part b   # Submit batch (2 min, then close laptop)")
        print("  python3 main_run.py --part c   # Download results (run next morning)")
        sys.exit(0)

    part = sys.argv[2].lower()
    if part == "a":
        run_part_a()
    elif part == "b":
        run_part_b()
    elif part == "c":
        run_part_c()
    else:
        print(f"Unknown part: {part}. Use a, b, or c.")
